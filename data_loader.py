"""
data_loader.py
──────────────
Pure-function file readers for BPS trade CSVs and the Neraca Komoditas xlsx.

All functions here are stateless. They are invoked once by `Problem.load()`
to populate the immutable Problem dataclass.

Files expected in the same directory:
  - Impor Kedelai Berdasarkan Negara Asal.csv
  - Volume Impor KEDELAI menurut pelabuhan utama.csv
  - Volume Impor Menurut Pelabuhan Utama (Berat bersih_ ribu ton), 2017-2024.csv
  - Proyeksi Neraca Komoditas Kedelai.xlsx              (optional — for demand)
"""
from __future__ import annotations

import csv
import os
import numpy as np


# ═══════════════════════════════════════════════════════════════════════════
#  Static name → index mappings
# ═══════════════════════════════════════════════════════════════════════════

BULAN_IDX = {
    'Januari': 0, 'Februari': 1, 'Maret': 2, 'April': 3, 'Mei': 4, 'Juni': 5,
    'Juli': 6, 'Agustus': 7, 'September': 8, 'Oktober': 9, 'November': 10,
    'Desember': 11,
}

PROV_NAME_IDX = {
    'ACEH': 0, 'SUMATERA UTARA': 1, 'SUMATERA BARAT': 2, 'RIAU': 3,
    'JAMBI': 4, 'SUMATERA SELATAN': 5, 'BENGKULU': 6, 'LAMPUNG': 7,
    'BANGKA BELITUNG': 8, 'KEP. BANGKA BELITUNG': 8,
    'KEPULAUAN RIAU': 9, 'DKI JAKARTA': 10, 'DKI JAKATA': 10, 'JAWA BARAT': 11,
    'JAWA TENGAH': 12, 'DI YOGYAKARTA': 13, 'JAWA TIMUR': 14,
    'BANTEN': 15, 'BALI': 16, 'NUSA TENGGARA BARAT': 17,
    'NUSA TENGGARA TIMUR': 18, 'KALIMANTAN BARAT': 19,
    'KALIMANTAN TENGAH': 20, 'KALIMANTAN SELATAN': 21,
    'KALIMANTAN TIMUR': 22, 'KALIMANTAN UTARA': 23,
    'SULAWESI UTARA': 24, 'SULAWESI TENGAH': 25, 'SULAWESI SELATAN': 26,
    'SULAWESI TENGGARA': 27, 'GORONTALO': 28, 'SULAWESI BARAT': 29,
    'MALUKU': 30, 'MALUKU UTARA': 31, 'PAPUA BARAT': 32,
    'PAPUA BARAT DAYA': 33, 'PAPUA': 34, 'PAPUA SELATAN': 35,
    'PAPUA TENGAH': 36, 'PAPUA PEGUNUNGAN': 37,
}

# Mapping port array index → CSV name (for soybean port volumes)
PORT_SOY_MAP = {
    0: 'BELAWAN',  1: 'TELUK BAYUR',  2: 'PANJANG',
    3: 'TANJUNG BALAI KARIMUN',  4: 'BATU AMPAR',  5: 'TANJUNG PRIOK',
    6: 'TANJUNG EMAS',  7: 'GRESIK',  8: 'TANJUNG PERAK',
    9: 'CIGADING',  10: 'SUPADIO (U)',
}

# Mapping port array index → annual CSV name (all-commodity throughput)
PORT_ANNUAL_CSV_MAP = {
    0: 'Belawan',  2: 'Panjang (Lampung, Sumatra)',
    3: 'Tanjung Balai Karimun',  4: 'Batu Ampar',  5: 'Tanjung Priok',
    6: 'Tanjung Emas',  8: 'Tanjung Perak',  9: 'Cigading',
}

# Defaults for ports without direct CSV entries (estimasi regional)
PORT_ANNUAL_DEFAULTS = {
    1: 3_000.0,    # Teluk Bayur (estimasi ~25% Sumatera Lainnya)
    7: 12_000.0,   # Gresik (estimasi dari Jawa Lainnya 29.517,6 kt)
    10: 800.0,     # Supadio (estimasi dari Kalimantan Lainnya 3.681,3 kt)
}

# Strategic multi-year granular registry. Airport rows are excluded entirely.
GRANULAR_IMPORT_COUNTRY_MAP = {
    'UNITED STATES': 0,
    'UNITED STATES OF AMERICA': 0,
    'CANADA': 1,
    'ARGENTINA': 2,
    'BRAZIL': 3,
    'MALAYSIA': 4,
    'BOLIVIA': 5,
    'URUGUAY': 6,
}

GRANULAR_IMPORT_COUNTRY_LABELS = [
    'USA', 'Kanada', 'Argentina', 'Brazil', 'Malaysia', 'Bolivia', 'Uruguay',
]

GRANULAR_PORT_MAP = {
    'BELAWAN': 0,
    'PANJANG': 1,
    'BATU AMPAR': 2,
    'TANJUNG PRIOK': 3,
    'TANJUNG EMAS': 4,
    'GRESIK': 5,
    'TANJUNG PERAK': 6,
    'CIGADING': 7,
    'CILACAP': 8,
    'TANJUNG INTAN': 8,
    'BANYUWANGI': 9,
    'TANJUNG WANGI': 9,
    'PONTIANAK': 10,
    'PELABUHAN PONTIANAK': 10,
}

GRANULAR_PORT_LABELS = [
    'Belawan', 'Panjang', 'Batu Ampar', 'Tanjung Priok', 'Tanjung Emas',
    'Gresik', 'Tanjung Perak', 'Cigading', 'Cilacap', 'Banyuwangi',
    'Pontianak',
]


# ═══════════════════════════════════════════════════════════════════════════
#  Helpers
# ═══════════════════════════════════════════════════════════════════════════

def parse_num(s) -> float:
    """Parse a number string that may have comma thousand-separators."""
    s = str(s).strip().strip('"')
    if not s or s == '-':
        return 0.0
    return float(s.replace(',', ''))


MONTH_KEYWORDS = {
    0: ['JANUARI', 'JAN'], 1: ['FEBRUARI', 'FEB'],
    2: ['MARET', 'MAR'], 3: ['APRIL', 'APR'],
    4: ['MEI', 'MAY'], 5: ['JUNI', 'JUN'],
    6: ['JULI', 'JUL'], 7: ['AGUSTUS', 'AGU', 'AUG'],
    8: ['SEPTEMBER', 'SEP'], 9: ['OKTOBER', 'OKT', 'OCT'],
    10: ['NOVEMBER', 'NOV'], 11: ['DESEMBER', 'DES', 'DEC'],
}


def _norm_text(value) -> str:
    if value is None:
        return ''
    return ' '.join(str(value).strip().upper().split())


def _parse_month(label) -> int | None:
    s = _norm_text(label)
    if not s:
        return None
    if s.startswith('['):
        s = s.split(']', 1)[-1].strip()
    for idx, kws in MONTH_KEYWORDS.items():
        if any(kw in s for kw in kws):
            return idx
    return None


def _backtrack_header(row, col: int) -> str:
    for i in range(col, 1, -1):
        value = _norm_text(row[i] if i < len(row) else None)
        if value:
            return value
    return ''


def _is_airport_port(port: str) -> bool:
    port = _norm_text(port)
    return (
        '(U)' in port
        or 'SOEKARNO' in port
        or 'JUANDA' in port
        or 'NGURAH RAI' in port
    )


def _load_workbook_rows(path: str) -> list[tuple]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to read granular import workbooks"
        ) from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 9:
        raise ValueError(f"Granular workbook is too short: {path}")
    return rows


def _iter_granular_cells(rows: list[tuple], years: tuple[int, ...]):
    row_countries = rows[1]
    row_ports = rows[2]
    row_months = rows[3]
    selected_years = set(years)

    for data_row in rows[5:]:
        if len(data_row) < 2 or data_row[1] is None:
            continue
        try:
            year = int(data_row[1])
        except (TypeError, ValueError):
            continue
        if year not in selected_years:
            continue

        for col in range(2, len(row_countries)):
            country = _backtrack_header(row_countries, col)
            port = _backtrack_header(row_ports, col)
            month = _parse_month(_backtrack_header(row_months, col))
            if not country or not port or month is None:
                continue
            if country.startswith('TOTAL') or country.startswith('NEGARA/'):
                continue

            raw_value = data_row[col] if col < len(data_row) else None
            if raw_value is None:
                value = 0.0
            else:
                value = float(raw_value)
            yield year, country, port, month, value


# ═══════════════════════════════════════════════════════════════════════════
#  CSV readers
# ═══════════════════════════════════════════════════════════════════════════

def read_bps_trade_csv(filename: str, data_dir: str) -> dict:
    """
    Read a BPS trade CSV (Impor Berdasarkan Negara / Pelabuhan).

    Format:
      Row 1: title
      Row 2: month headers
      Row 3: Vol/Val subheaders
      Row 4+: Name, Jan_Vol, Jan_Val, Feb_Vol, Feb_Val, ..., Tot_Vol, Tot_Val

    Returns:
      {NAME (upper) → (monthly_volumes_kg[12], monthly_values_usd[12])}

    Duplicate names (e.g. two CIGADING rows) are summed automatically.
    """
    path = os.path.join(data_dir, filename)
    with open(path, encoding='utf-8') as f:
        rows = list(csv.reader(f))
    data: dict = {}
    for row in rows[3:]:
        if not row or not row[0].strip():
            continue
        name = row[0].strip().upper()
        vols = np.array([parse_num(row[i]) if i < len(row) else 0.0
                         for i in range(1, 25, 2)])   # 1,3,...,23 → 12 months
        vals = np.array([parse_num(row[i]) if i < len(row) else 0.0
                         for i in range(2, 26, 2)])   # 2,4,...,24 → 12 months
        if name in data:
            data[name] = (data[name][0] + vols, data[name][1] + vals)
        else:
            data[name] = (vols, vals)
    return data


def read_port_annual_csv(data_dir: str) -> dict:
    """
    Read 'Volume Impor Menurut Pelabuhan Utama (Berat bersih_ ribu ton), 2017-2024.csv'.

    Returns: {port_name → 2024 annual throughput (ribu ton, all commodities)}
    """
    path = os.path.join(data_dir,
        "Volume Impor Menurut Pelabuhan Utama (Berat bersih_ ribu ton), 2017-2024.csv")
    with open(path, encoding='utf-8') as f:
        rows = list(csv.reader(f))
    data: dict = {}
    for row in rows[3:]:
        if len(row) < 10:
            continue
        port = row[1].strip()
        if not port:
            continue
        val_str = row[9].strip() if len(row) > 9 else ''
        if not val_str:
            continue
        data[port] = parse_num(val_str)
    return data


def read_granular_import_pipeline(
    data_dir: str,
    n_imp: int = 7,
    n_port: int = 11,
    n_period: int = 12,
    years: tuple[int, ...] = (2022, 2023, 2024, 2025),
    initial_year: int = 2024,
    exch_rate: float = 16_000.0,
    import_cap_buffer: float = 1.01,
    port_cap_buffer: float = 1.1,
) -> dict:
    """
    Build the current import data pipeline from granular BPS workbooks.

    Sources:
      - bobot workbook: import volume by country-port-month-year, in kg
      - nilai dolar workbook: import value by country-port-month-year, in USD

    Policy:
      - airport rows are excluded entirely
      - selected strategic countries seed source capacity over 2022-2025
      - initial imports remain realized 2024 flows for selected ports only
      - port capacity uses all-country non-airport evidence for selected ports

    Buffer rationale:
      - import_cap_buffer = 1.01: the max historical monthly volume over a
        finite 4-year window may slightly underestimate the true population
        maximum; a 1 % sampling margin avoids infeasibility from sampling
        error.  Global soybean supply is structurally concentrated (Brazil
        40 %, USA 28 %, Argentina 12 %), so source capacity is inherently
        tight and large multipliers are unwarranted.
      - port_cap_buffer = 1.1: targets ~90% utilisation consistent with
        Pelindo-reported rates for major soybean ports (Tanjung Priok 90%,
        Tanjung Emas 95%, Tanjung Perak 87%); 1/0.90 ≈ 1.11 rounded to 1.1.
    """
    if n_imp != len(GRANULAR_IMPORT_COUNTRY_LABELS):
        raise ValueError(
            f"granular pipeline expects {len(GRANULAR_IMPORT_COUNTRY_LABELS)} "
            f"import countries, got {n_imp}"
        )
    if n_port != len(GRANULAR_PORT_LABELS):
        raise ValueError(
            f"granular pipeline expects {len(GRANULAR_PORT_LABELS)} ports, got {n_port}"
        )

    weight_path = os.path.join(
        data_dir, "Impor kedelai dan negara asal dan tujuan berdasarkan bobot.xlsx"
    )
    value_path = os.path.join(
        data_dir, "Impor kedelai dan negara asal dan tujuan berdasarkan nilai dolar.xlsx"
    )
    if not os.path.exists(weight_path):
        raise FileNotFoundError(weight_path)
    if not os.path.exists(value_path):
        raise FileNotFoundError(value_path)

    year_idx = {year: idx for idx, year in enumerate(years)}
    hist_import = np.zeros((n_imp, n_port, n_period), dtype=float)
    country_year_month_kg = np.zeros((len(years), n_imp, n_period), dtype=float)
    port_year_month_kg = np.zeros((len(years), n_port, n_period), dtype=float)
    country_total_weight_kg = np.zeros(n_imp, dtype=float)
    country_total_value_usd = np.zeros(n_imp, dtype=float)
    airport_weight_kg = 0.0
    airport_value_usd = 0.0

    weight_rows = _load_workbook_rows(weight_path)
    for year, country, port, month, weight_kg in _iter_granular_cells(weight_rows, years):
        if _is_airport_port(port):
            airport_weight_kg += weight_kg
            continue

        s_idx = GRANULAR_IMPORT_COUNTRY_MAP.get(country)
        h_idx = GRANULAR_PORT_MAP.get(port)
        y_idx = year_idx[year]

        if s_idx is not None:
            country_year_month_kg[y_idx, s_idx, month] += weight_kg
            country_total_weight_kg[s_idx] += weight_kg
            if year == initial_year and h_idx is not None:
                hist_import[s_idx, h_idx, month] += weight_kg / 1000.0

        if h_idx is not None:
            port_year_month_kg[y_idx, h_idx, month] += weight_kg

    value_rows = _load_workbook_rows(value_path)
    for _, country, port, _, value_usd in _iter_granular_cells(value_rows, years):
        if _is_airport_port(port):
            airport_value_usd += value_usd
            continue

        s_idx = GRANULAR_IMPORT_COUNTRY_MAP.get(country)
        if s_idx is not None:
            country_total_value_usd[s_idx] += value_usd

    imp_cap_normal = country_year_month_kg.max(axis=0) / 1000.0 * import_cap_buffer
    port_thru_cap = port_year_month_kg.max(axis=0) / 1000.0 * port_cap_buffer

    c_purch = np.empty(n_imp, dtype=float)
    for s in range(n_imp):
        if country_total_weight_kg[s] > 0 and country_total_value_usd[s] > 0:
            c_purch[s] = (
                country_total_value_usd[s] / country_total_weight_kg[s]
                * 1000.0 * exch_rate
            )
        else:
            c_purch[s] = 7_000_000.0

    summary = "  ".join(
        f"{name}={hist_import[idx].sum():,.0f}"
        for idx, name in enumerate(GRANULAR_IMPORT_COUNTRY_LABELS)
    )
    cap_summary = "  ".join(
        f"{name}={imp_cap_normal[idx].sum():,.0f}"
        for idx, name in enumerate(GRANULAR_IMPORT_COUNTRY_LABELS)
    )
    print(
        f"[DATA] Granular import initial (year={initial_year}, no-airport): "
        f"{summary}  Total={hist_import.sum():,.0f} ton"
    )
    print(
        f"[DATA] Strategic import capacity ({min(years)}-{max(years)}, "
        f"buffer={import_cap_buffer:g}): {cap_summary} ton/month-sum"
    )
    print(
        f"[DATA] Granular port capacity ({min(years)}-{max(years)}, "
        f"buffer={port_cap_buffer:g}): Total={port_thru_cap.sum():,.0f} "
        f"ton/month-sum; airport excluded={airport_weight_kg / 1000.0:,.0f} ton"
    )

    return {
        "hist_import": hist_import,
        "imp_cap_normal": imp_cap_normal,
        "port_thru_cap": port_thru_cap,
        "c_purch": c_purch,
        "diagnostics": {
            "years": tuple(years),
            "initial_year": initial_year,
            "airport_weight_excluded_kg": airport_weight_kg,
            "airport_value_excluded_usd": airport_value_usd,
            "country_total_weight_kg": country_total_weight_kg,
            "country_total_value_usd": country_total_value_usd,
            "selected_country_labels": list(GRANULAR_IMPORT_COUNTRY_LABELS),
            "selected_port_labels": list(GRANULAR_PORT_LABELS),
        },
    }


def read_disaggregated_import(data_dir: str, n_imp: int = 3,
                              n_port: int = 11, n_period: int = 12,
                              year: int = 2024) -> np.ndarray | None:
    """
    Read 'Impor kedelai dan negara asal dan tujuan berdasarkan bobot.xlsx'.

    Returns:
      (n_imp, n_port, n_period) ndarray of import volume in TONS,
      or None if file/library missing.

    The file has a sparse multi-level header:
      Row 1: Country names (merged across columns)
      Row 2: Port names (merged across columns)
      Row 3: Month labels (merged across columns)
      Rows 5-8: Data for years 2022-2025

    Non-model ports (Cilacap, Banyuwangi, etc.) have their volume
    redistributed to the nearest model port by geographic cluster.
    """
    path = os.path.join(data_dir,
        "Impor kedelai dan negara asal dan tujuan berdasarkan bobot.xlsx")
    if not os.path.exists(path):
        return None
    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 9:
        return None

    row_countries = rows[1]
    row_ports = rows[2]
    row_months = rows[3]
    year_row_map = {2022: 5, 2023: 6, 2024: 7, 2025: 8}
    data_row_idx = year_row_map.get(year, 7)
    if data_row_idx >= len(rows):
        return None

    IMP_FILE_TO_MODEL = {
        'UNITED STATES': 0, 'CANADA': 1, 'BRAZIL': 2,
    }

    PORT_FILE_TO_MODEL = {
        'BELAWAN': 0, 'TELUK BAYUR': 1, 'PANJANG': 2,
        'TANJUNG BALAI KARIMUN': 3, 'BATU AMPAR': 4,
        'TANJUNG PRIOK': 5, 'TANJUNG EMAS': 6, 'GRESIK': 7,
        'TANJUNG PERAK': 8, 'CIGADING': 9, 'SUPADIO': 10,
    }

    PORT_ALT_TO_MODEL = {
        'SOEKARNO-HATTA (U)': 10, 'SOEKARNO-HATTA': 10,
        'JUANDA (U)-SURABAYA': 8, 'JUANDA': 8,
        'SUPADIO (U)': 10, 'CILACAP': 6,
        'BANYUWANGI': 8, 'PONTIANAK': 10,
        'TANJUNG BERINGIN': 0, 'MERAK': 9,
        'NGURAH RAI (U)': 8, 'JOHOR (U)': 4,
        'JKT./ POS PASAR BARU': 5, 'BATU AMPAR': 4,
    }
    PORT_FILE_TO_MODEL.update(PORT_ALT_TO_MODEL)

    MONTH_KEYWORDS = {
        0: ['JANUARI', 'JAN'], 1: ['FEBRUARI', 'FEB'],
        2: ['MARET', 'MAR'], 3: ['APRIL', 'APR'],
        4: ['MEI', 'MAY'], 5: ['JUNI', 'JUN'],
        6: ['JULI', 'JUL'], 7: ['AGUSTUS', 'AGU', 'AUG'],
        8: ['SEPTEMBER', 'SEP'], 9: ['OKTOBER', 'OKT', 'OCT'],
        10: ['NOVEMBER', 'NOV'], 11: ['DESEMBER', 'DES', 'DEC'],
    }

    def _parse_month(label):
        if label is None:
            return None
        s = str(label).strip().upper()
        if s.startswith('['):
            s = s.split(']')[-1].strip()
        for idx, kws in MONTH_KEYWORDS.items():
            for kw in kws:
                if kw in s:
                    return idx
        return None

    def _backtrack(row, col):
        for i in range(col, 1, -1):
            if row[i] is not None:
                return str(row[i]).strip().upper()
        return None

    x_imp = np.zeros((n_imp, n_port, n_period))
    overflow = np.zeros(n_period)

    for col in range(2, len(row_countries)):
        country = _backtrack(row_countries, col)
        port = _backtrack(row_ports, col)
        month = _parse_month(_backtrack(row_months, col))

        if port and ('(U)' in port or 'SOEKARNO' in port) and 'SUPADIO' not in port:
            continue

        if country is None or port is None or month is None:
            continue

        s_idx = IMP_FILE_TO_MODEL.get(country)
        if s_idx is None:
            continue

        val = rows[data_row_idx][col] if col < len(rows[data_row_idx]) else None
        if val is None:
            overflow[month] += 0.0
            continue
        val_ton = float(val) / 1000.0

        h_idx = PORT_FILE_TO_MODEL.get(port)
        if h_idx is not None:
            x_imp[s_idx, h_idx, month] += val_ton
        else:
            overflow[month] += val_ton

    for t in range(n_period):
        if overflow[t] > 0:
            port_totals = x_imp[:, :, t].sum(axis=0)
            total_existent = port_totals[port_totals > 0].sum()
            if total_existent > 0:
                for h in range(n_port):
                    if port_totals[h] > 0:
                        share = port_totals[h] / total_existent
                        x_imp[:, h, t] += share * overflow[t] * (x_imp[:, h, t] / max(port_totals[h], 1e-9))

    parts = [f"S{s}={x_imp[s].sum():,.0f}" for s in range(x_imp.shape[0])]
    print(f"[DATA] Disaggregated import (year={year}): "
          f"{'  '.join(parts)}  Total={x_imp.sum():,.0f} ton")
    return x_imp


def read_province_import(data_dir: str, n_prov: int = 38,
                         n_period: int = 12,
                         year: int = 2024) -> np.ndarray | None:
    """
    Read 'Ekspor Impor Kedelai.xlsx' → province-level monthly volume.

    PERHATIAN: File ini mencatat PERDAGANGAN ANTAR-PROVINSI (bukan impor
    internasional). Angka di sini menggambarkan berapa ton kedelai yang
    diperdagangkan antar-provinsi, bukan berapa yang diimpor dari USA/Kanada.
    Fungsi ini disimpan untuk keperluan analisis/referensi saja dan TIDAK
    digunakan dalam build_historical_initial() atau pembentukan initial solution.

    Returns:
      (n_prov, n_period) ndarray of inter-provincial trade in TONS,
      or None if file/library missing.
    """
    path = os.path.join(data_dir, "Ekspor Impor Kedelai.xlsx")
    if not os.path.exists(path):
        return None
    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['data']

    BULAN_IDX = {
        'JANUARI': 0, 'FEBRUARI': 1, 'MARET': 2, 'APRIL': 3,
        'MEI': 4, 'JUNI': 5, 'JULI': 6, 'AGUSTUS': 7,
        'SEPTEMBER': 8, 'OKTOBER': 9, 'NOVEMBER': 10, 'DESEMBER': 11,
    }

    x_prov = np.zeros((n_prov, n_period))

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 8:
            continue
        prov = str(row[1]).strip().upper() if row[1] else ''
        bulan = str(row[4]).strip().upper() if row[4] else ''
        tahun = row[5]
        imp = float(row[6]) if row[6] is not None else 0.0

        if tahun != year:
            continue

        i = PROV_NAME_IDX.get(prov)
        t = BULAN_IDX.get(bulan)
        if i is not None and t is not None and i < n_prov and t < n_period:
            x_prov[i, t] += imp

    wb.close()

    for i in range(n_prov):
        if x_prov[i].sum() == 0:
            pass

    print(f"[DATA] Province import (year={year}): total={x_prov.sum():,.0f} ton, "
          f"provinces with data={np.count_nonzero(x_prov.sum(axis=1) > 0)}")
    return x_prov


def load_neraca_demand(data_dir: str, n_prov: int, n_period: int,
                       year: int = 2024) -> np.ndarray | None:
    """
    Load monthly demand per province from 'Proyeksi Neraca Komoditas Kedelai.xlsx'.
    Column 'total_kebutuhan (ton)' is D_{i,t} directly.

    Returns:
      (n_prov, n_period) ndarray of demand in tons, or None if file/library missing.

    Some provinces are missing in some years (DKI Jakarta only in 2025,
    Papua Pegunungan only in 2024). To fill gaps, we load every year and
    fall back to the nearest available year for any (prov, month) cell that
    is missing in `year`.
    """
    path = os.path.join(data_dir, "Proyeksi Neraca Komoditas Kedelai.xlsx")
    if not os.path.exists(path):
        return None
    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['data']

    # Load all years into a dict, prioritise `year` later
    all_data: dict = {}     # (year, prov_idx, month_idx) → kebutuhan
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 8:
            continue
        prov  = str(row[1]).strip().upper() if row[1] else ''
        bulan = str(row[4]).strip()         if row[4] else ''
        tahun = int(row[5])                  if row[5] else 0
        kebutuhan = float(row[7]) if row[7] is not None else 0.0
        i = PROV_NAME_IDX.get(prov)
        t = BULAN_IDX.get(bulan)
        if i is not None and t is not None and i < n_prov and t < n_period:
            all_data[(tahun, i, t)] = max(kebutuhan, 0.0)
    wb.close()

    # Build the array, fall back to nearest year for missing cells
    demand = np.full((n_prov, n_period), np.nan)
    fallback_years = [year - 1, year + 1, year - 2, year + 2]
    for i in range(n_prov):
        for t in range(n_period):
            val = all_data.get((year, i, t))
            if val is not None:
                demand[i, t] = val
                continue
            for fy in fallback_years:
                val = all_data.get((fy, i, t))
                if val is not None:
                    demand[i, t] = val
                    break
    return demand
